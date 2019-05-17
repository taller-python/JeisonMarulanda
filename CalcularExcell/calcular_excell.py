"""Taller python"""
from openpyxl import load_workbook

FILE_PATH = 'operaciones.xlsx'
SHEET = 'SUMA'

WORKBOOK = load_workbook(FILE_PATH, read_only=True)
SH = WORKBOOK[SHEET]

SUMA = open('SUMA.txt', 'w')
for num1, num2 in SH.iter_rows(min_row=2):
    try:
        resultado = int(num1.value+num2.value)
    except:
        resultado = 'Error'
    SUMA.write(str(resultado)+"\n")
SUMA.close()

SHEET = 'RESTA'
SH = WORKBOOK[SHEET]

RESTA = open('RESTA.txt', 'w')
for num1, num2 in SH.iter_rows(min_row=2):
    try:
        resultado = int(num1.value-num2.value)
    except:
        resultado = 'Error'
    RESTA.write(str(resultado)+"\n")
RESTA.close()

SHEET = 'MULTIPLICACIÓN'
SH = WORKBOOK[SHEET]

MULTI = open('MULTIPLICACION.txt', 'w')
for num1, num2 in SH.iter_rows(min_row=2):
    try:
        resultado = int(num1.value*num2.value)
    except:
        resultado = 'Error'
    MULTI.write(str(resultado)+"\n")
MULTI.close()

SHEET = 'DIVISIÓN'
SH = WORKBOOK[SHEET]

DIVI = open('DIVISION.txt', 'w')
for num1, num2 in SH.iter_rows(min_row=2):
    try:
        resultado = int(num1.value/num2.value)
    except:
        resultado = 'Error'
    DIVI.write(str(resultado)+"\n")
DIVI.close()

print('OK')
